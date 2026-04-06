from datetime import datetime
from openpyxl import load_workbook

DATE_FIELDS = {4, 5, 6, 7, 17, 18, 19}


def load_students(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    field_numbers = rows[1]

    students = []
    for row in rows[2:]:
        if not any(row):
            continue
        student = {}
        for col_idx, field_num in enumerate(field_numbers):
            if not isinstance(field_num, (int, float)):
                continue
            field_num = int(field_num)
            value = row[col_idx]
            if value is None:
                value = ""
            if isinstance(value, datetime) or field_num in DATE_FIELDS:
                try:
                    value = value.strftime("%d.%m.%Y")
                except AttributeError:
                    pass
            student[field_num] = str(value).strip()
        students.append(student)

    wb.close()
    return students